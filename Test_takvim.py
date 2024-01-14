from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait #ilgili driverı bekleten yapı
from selenium.webdriver.support import expected_conditions as ec #beklenen koşullar
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import pytest
import openpyxl
from constants import globalConstants as c


class Test_tobetoPlatformLogin():
    def setup_method(self):
        self.driver = webdriver.Chrome()
        self.driver.get(c.BASE_URL)
        self.driver.maximize_window()

    def teardown_metahod(self):
        self.driver.quit()
    
    def getData():
        excel = openpyxl.load_workbook("data/invalidLogin.xlsx")
        sheet = excel["Sheet1"]
        rows = sheet.max_row
        data = []
        for i in range(2,rows+1):
            email = sheet.cell(i,1).value
            password = sheet.cell(i,2).value
            data.append((email,password))

        return data
    
    
    #1)Giriş yap alanı görüntülenebilir ve işlevselliği test edilecektir.
    def test_visibility_of_login_page(self):
        takvim = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='__next']/div/section[1]/div[2]/div")))
        takvim.click()
        takvim_page_title = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH,"/html/body/div[4]/div/div/div[1]/span")))
        assert takvim_page_title.text == "Eğitim ve Etkinlik Takvimi", "Başlık 'Eğitim ve Etkinlik Takvimi' olarak bekleniyor."
        assert takvim.is_displayed(), "Takvim simgesi görüntülenmiyor."
        #aramacubuğu kontrolü
        aramacubugu=WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='search-event']")))
        aramacubugu.click()
        arama_metni = "yazılım kalite"
        aramacubugu.send_keys(arama_metni)
        #eğitmenaramayapılacak
        
        #egitmen=WebDriverWait(self.driver,6).until(ec.visibility_of_element_located((By.XPATH,"/html/body/div[3]/div/div/div[2]/div/div/div[1]/div/div[2]/div[2]/div/div[1]/div[2]")))
        #egitmen.click()
        #listbox_element = WebDriverWait(self.driver, 6).until(ec.visibility_of_element_located((By.XPATH, "/html/body/div[3]/div/div/div[2]/div/div/div[1]/div/div[2]/div[2]/div/div[1]/div[2]")))
        #listbox = Select(listbox_element)
        #listbox.select_by_index(0)
        
        #eğitim durumu
        #bitmisdersler=WebDriverWait(self.driver,6).until(ec.visibility_of_element_located((By.XPATH,"/html/body/div[3]/div/div/div[2]/div/div/div[1]/div/div[3]/div[2]/span[1]/input")))
        #bitmisdersler.click()
        #bitmisdersler = WebDriverWait(self.driver, 10).until(ec.presence_of_element_located((By.XPATH, "/html/body/div[3]/div/div/div[2]/div/div/div[1]/div/div[3]/div[2]/span[1]/input")))
        #bitmisdersler.click()
        # Label elementini bulma
        #label_element = self.driver.find_element_by_xpath("/html/body/div[3]/div/div/div[2]/div/div/div[1]/div/div[3]/div[2]/span[1]/input, 'Bitmiş Dersler')]")
# Label elementine tıklama
        #label_element.click()

